import pymysql
import pandas as pd
from openpyxl import load_workbook

DB_CONFIG = {
  'host': 'localhost',
  'user': 'root',
  'password': 'password',
  'database': 'retail_sales'
}

def load_data_from_sql(db_config, table_name):
  connection = pymysql.connect(**db_config)
  query = f"SELECT * FROM {table_name}"
  df = pd.read_sql_query(query, connection)
  connection.close()
  return df

df = load_data_from_sql(DB_CONFIG, 'transformed_sales')
df.head(5)

def append_raw_data():
  book = load_workbook('./raw/online_retial.xlsx')
  
  with pd.ExcelWriter('./data/online_retail.xlsx', engine='openpyxl', mode = 'a') as writer:
    writer.book = book
    writer.sheets = {ws.title: ws for ws in book.worksheets}

    df.to_excel(writer, sheet_name='Online Retail', index = False,
                    header=False, startrow=writer.sheets['Online Retail'].max_row)

    print('Data appended successfully!')

append_raw_data